# --------------------------------------------------------------
# ReconRaptor — Vendor Reconciliation (FINAL • Tier de-dup • FIXED)
# --------------------------------------------------------------
import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from difflib import SequenceMatcher
import numpy as np

# ==================== PAGE CONFIG & CSS ======================
st.set_page_config(page_title="ReconRaptor — Vendor Reconciliation", layout="wide")
st.markdown(
    """
<style>
.big-title {
    font-size: 3rem !important;
    font-weight: 700;
    text-align: center;
    background: linear-gradient(90deg, #1E88E5, #42A5F5);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    margin-bottom: 1rem;
}
.section-title {
    font-size: 1.8rem !important;
    font-weight: 600;
    color: #1565C0;
    border-bottom: 2px solid #42A5F5;
    padding-bottom: 0.5rem;
    margin-top: 2rem;
}
.metric-container {
    padding: 1.2rem !important;
    border-radius: 12px !important;
    margin-bottom: 1rem;
    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
}
.perfect-match   {background:#2E7D32;color:#fff;font-weight:bold;}
.difference-match{background:#FF8F00;color:#fff;font-weight:bold;}
.tier2-match     {background:#26A69A;color:#fff;font-weight:bold;}
.tier3-match     {background:#7E57C2;color:#fff;font-weight:bold;}
.missing-erp     {background:#C62828;color:#fff;font-weight:bold;}
.missing-vendor  {background:#AD1457;color:#fff;font-weight:bold;}
.payment-match   {background:#004D40;color:#fff;font-weight:bold;}
</style>
""",
    unsafe_allow_html=True,
)

st.markdown('<h1 class="big-title">ReconRaptor</h1>', unsafe_allow_html=True)
st.markdown("<p style='text-align: center; font-size: 1.3rem; color: #555;'>Intelligent Vendor Invoice Reconciliation</p>", unsafe_allow_html=True)

# ====================== HELPERS ==========================
def fuzzy_ratio(a, b):
    return SequenceMatcher(None, str(a), str(b)).ratio()

def normalize_number(v):
    if pd.isna(v) or str(v).strip() == "":
        return 0.0
    s = re.sub(r"[^\d,.\-]", "", str(v).strip())
    if s.count(",") == 1 and s.count(".") == 1:
        if s.find(",") > s.find("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif s.count(",") == 1:
        s = s.replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "", s.count(".") - 1)
    try:
        return float(s)
    except:
        return 0.0

def normalize_date(v):
    if pd.isna(v) or str(v).strip() == "":
        return ""
    s = str(v).strip().replace(".", "/").replace("-", "/").replace(",", "/")
    for fmt in [
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%m/%d/%Y", "%m-%d-%Y",
        "%Y/%m/%d", "%Y-%m-%d",
        "%d/%m/%y", "%d-%m-%y", "%d.%m.%y",
        "%m/%d/%y", "%m-%d-%y",
        "%Y.%m.%d",
    ]:
        try:
            d = pd.to_datetime(s, format=fmt, errors="coerce")
            if not pd.isna(d):
                return d.strftime("%Y-%m-%d")
        except:
            continue
    d = pd.to_datetime(s, errors="coerce")
    if pd.isna(d):
        d = pd.to_datetime(s, errors="coerce")
    return d.strftime("%Y-%m-%d") if not pd.isna(d) else ""

def clean_invoice_code(v):
    if not v:
        return ""
    s = str(v).strip().lower()
    parts = re.split(r"[-_.\s]", s)
    for p in reversed(parts):
        if re.fullmatch(r"\d{1,}", p) and not re.fullmatch(r"20[0-3]\d", p):
            s = p.lstrip("0")
            break
    s = re.sub(r"^(αρ|τιμ|pf|ab|inv|tim|cn|ar|pa|πφ|πα|apo|ref|doc|num|no|apd|vs)\W*", "", s)
    s = re.sub(r"20\d{2}", "", s)
    s = re.sub(r"[^a-z0-9]", "", s)
    s = re.sub(r"^0+", "", s)
    s = re.sub(r"[^\d]", "", s)
    return s or "0"

def normalize_columns(df, tag):
    mapping = {
        "invoice": [
            "invoice", "invoice number", "inv no", "factura", "fact", "nº", "num",
            "numero", "document", "doc", "ref", "αρ", "παραστ", 
            "alternative document", "alt document", "alt. document", "alternative doc"
        ],
        "credit":  ["credit", "haber", "credito", "abono"],
        "debit":   ["debit", "debe", "cargo", "importe", "amount", "valor", "total"],
        "reason":  ["reason", "motivo", "concepto", "descripcion", "detalle", "περιγραφή"],
        "date":    ["date", "fecha", "data", "issue date", "posting date", "ημερομηνία"]
    }

    rename_map = {}
    cols_lower = {c: str(c).strip().lower() for c in df.columns}

    for key, aliases in mapping.items():
        for col, low in cols_lower.items():
            if any(a in low for a in aliases):
                rename_map[col] = f"{key}_{tag}"

    out = df.rename(columns=rename_map)

    # Guarantee debit/credit even if missing
    for req in ["debit", "credit"]:
        c = f"{req}_{tag}"
        if c not in out.columns:
            out[c] = 0.0

    # Normalize date column
    if f"date_{tag}" in out.columns:
        out[f"date_{tag}"] = out[f"date_{tag}"].apply(normalize_date)

    st.write(f"✅ Normalized {tag.upper()} columns:", list(out.columns))
    return out

def style(df, css):
    return df.style.apply(lambda _: [css] * len(_), axis=1)

# ==================== MATCHING CORE ==========================
def match_invoices(erp_df, ven_df):
    # ---- classify invoice type: INV / CN / IGNORE ----
    def doc_type(row, tag):
        txt = (str(row.get(f"reason_{tag}", "")) + " " + str(row.get(f"invoice_{tag}", ""))).lower()
        debit  = normalize_number(row.get(f"debit_{tag}", 0))
        credit = normalize_number(row.get(f"credit_{tag}", 0))
        pay_kw = [
            "πληρωμ", "payment", "remittance", "bank transfer",
            "transferencia", "trf", "remesa", "pago", "deposit",
            "μεταφορά", "έμβασμα", "εξόφληση", "pagado", "paid", "cobro"
        ]
        if any(k in txt for k in pay_kw):
            return "IGNORE"
        if any(k in txt for k in ["credit", "nota", "abono", "cn", "πιστωτικό", "πίστωση", "ακυρωτικό"]):
            return "CN"
        if any(k in txt for k in ["factura", "invoice", "inv", "τιμολόγιο", "παραστατικό"]) or debit > 0 or credit > 0:
            return "INV"
        return "UNKNOWN"

    erp_df["__type"] = erp_df.apply(lambda r: doc_type(r, "erp"), axis=1)
    ven_df["__type"] = ven_df.apply(lambda r: doc_type(r, "ven"), axis=1)

    # 🚫 Exclude payments before consolidation
    erp_df = erp_df[erp_df["__type"].isin(["INV", "CN"])].copy()
    ven_df = ven_df[ven_df["__type"].isin(["INV", "CN"])].copy()

    # 🔹 Consolidate same invoice codes (INV + CN netting)
    # FIXED: Group by CLEANED invoice code so "INV 123" and "CN 123" both → "123"
    def consolidate(df, tag):
        if f"invoice_{tag}" not in df.columns:
            return df
        
        # Group by CLEANED invoice code so "INV 123" and "CN 123" both → "123"
        df = df.copy()
        df["__inv_clean"] = df[f"invoice_{tag}"].apply(clean_invoice_code)
        
        grouped = []
        for inv_clean, g in df.groupby("__inv_clean", dropna=False):
            if not inv_clean or inv_clean == "0":
                continue

            total = 0.0
            for _, r in g.iterrows():
                d = normalize_number(r.get(f"debit_{tag}", 0))
                c = normalize_number(r.get(f"credit_{tag}", 0))

                # Raw amount from ERP/Vendor: debit positive, credit negative
                raw = d - c

                if r.get("__type") == "CN":
                    # Credit notes must ALWAYS reduce the invoice
                    raw = -abs(raw)   # ensure negative
                else:
                    # Invoices must ALWAYS increase the balance
                    raw = abs(raw)    # ensure positive

                total += raw

            net_val = round(total, 2)

            # 🚫 Skip fully cancelling documents (INV + CN = 0 net)
            if abs(net_val) == 0:
                continue

            # Use the INV row as base (not CN), fallback to first row
            inv_rows = g[g["__type"] == "INV"]
            base = inv_rows.iloc[0].copy() if not inv_rows.empty else g.iloc[0].copy()
            # For matching & display we use absolute net amount
            base["__amt"] = abs(net_val)
            grouped.append(base)

        return pd.DataFrame(grouped)

    erp_df = consolidate(erp_df, "erp")
    ven_df = consolidate(ven_df, "ven")

    # --- Ensure __amt always exists ---
    if "__amt" not in erp_df.columns:
        erp_df["__amt"] = (
            erp_df.get("debit_erp", 0).apply(normalize_number)
            - erp_df.get("credit_erp", 0).apply(normalize_number)
        ).abs().round(2)

    if "__amt" not in ven_df.columns:
        ven_df["__amt"] = (
            ven_df.get("debit_ven", 0).apply(normalize_number)
            - ven_df.get("credit_ven", 0).apply(normalize_number)
        ).abs().round(2)

    # Normalize __amt (final cleanup)
    erp_df["__amt"] = erp_df["__amt"].apply(lambda x: round(normalize_number(x), 2))
    ven_df["__amt"] = ven_df["__amt"].apply(lambda x: round(normalize_number(x), 2))

    # 🔹 Exclude payments entirely (keep only invoices & credit notes)
    erp_use = erp_df[erp_df["__type"].isin(["INV", "CN"])].copy()
    ven_use = ven_df[ven_df["__type"].isin(["INV", "CN"])].copy()

    # ---------- Tier-1 exact matches ----------
    matched, used_vendor = [], set()
    for e_idx, e in erp_use.iterrows():
        e_inv = str(e.get("invoice_erp", "")).strip()
        e_amt = round(float(e.get("__amt", 0.0)), 2)

        for v_idx, v in ven_use.iterrows():
            if v_idx in used_vendor:
                continue
            v_inv = str(v.get("invoice_ven", "")).strip()
            v_amt = round(float(v.get("__amt", 0.0)), 2)

            if e_inv == v_inv:
                diff = abs(e_amt - v_amt)
                status = "Perfect Match" if diff <= 0.01 else "Difference Match"
                matched.append({
                    "ERP Invoice": e_inv,
                    "Vendor Invoice": v_inv,
                    "ERP Amount": e_amt,
                    "Vendor Amount": v_amt,
                    "Difference": round(diff, 2),
                    "Status": status
                })
                used_vendor.add(v_idx)
                break

    matched_df = pd.DataFrame(matched)

    # ---------- Remaining / missing ----------
    erp_use["__inv_norm"] = erp_use["invoice_erp"].apply(clean_invoice_code)
    ven_use["__inv_norm"] = ven_use["invoice_ven"].apply(clean_invoice_code)

    miss_erp = erp_use[~erp_use["__inv_norm"].isin(
        matched_df["ERP Invoice"].apply(clean_invoice_code) if not matched_df.empty else []
    )]
    miss_ven = ven_use[~ven_use["__inv_norm"].isin(
        matched_df["Vendor Invoice"].apply(clean_invoice_code) if not matched_df.empty else []
    )]

    miss_erp = miss_erp.rename(columns={"invoice_erp": "Invoice", "__amt": "Amount", "date_erp": "Date"})
    miss_ven = miss_ven.rename(columns={"invoice_ven": "Invoice", "__amt": "Amount", "date_ven": "Date"})
    keep_cols = ["Invoice", "Amount", "Date"]
    miss_erp = miss_erp[[c for c in keep_cols if c in miss_erp.columns]].reset_index(drop=True)
    miss_ven = miss_ven[[c for c in keep_cols if c in miss_ven.columns]].reset_index(drop=True)
    return matched_df, miss_erp, miss_ven

# ==================== TIERS 2 & 3 ==========================
# ------- Tier-2: fuzzy invoice + small amount tolerance -------
def tier2_match(erp_miss, ven_miss):
    if erp_miss.empty or ven_miss.empty:
        return pd.DataFrame(), set(), set(), erp_miss.copy(), ven_miss.copy()

    e = erp_miss.copy()
    v = ven_miss.copy()
    matches, used_e, used_v = [], set(), set()

    for ei, er in e.iterrows():
        if ei in used_e:
            continue
        e_inv = str(er.get("Invoice", ""))
        e_amt = round(float(er.get("Amount", 0.0)), 2)
        e_code = clean_invoice_code(e_inv)

        for vi, vr in v.iterrows():
            if vi in used_v:
                continue
            v_inv = str(vr.get("Invoice", ""))
            v_amt = round(float(vr.get("Amount", 0.0)), 2)
            v_code = clean_invoice_code(v_inv)

            diff = abs(e_amt - v_amt)
            sim = 1.0 if (e_code in v_code or v_code in e_code) else fuzzy_ratio(e_code, v_code)

            if diff <= 1.00 and sim >= 0.90:
                matches.append({
                    "ERP Invoice": e_inv,
                    "Vendor Invoice": v_inv,
                    "ERP Amount": e_amt,
                    "Vendor Amount": v_amt,
                    "Difference": round(diff, 2),
                    "Fuzzy Score": round(sim, 2),
                    "Match Type": "Tier-2"
                })
                used_e.add(ei)
                used_v.add(vi)
                break

    mdf = pd.DataFrame(matches)
    rem_e = e[~e.index.isin(used_e)].copy()
    rem_v = v[~v.index.isin(used_v)].copy()
    return mdf, used_e, used_v, rem_e, rem_v

# ------- Tier-3: same DATE + strong fuzzy (no amount threshold) -------
def tier3_match(erp_miss, ven_miss):
    if erp_miss.empty or ven_miss.empty:
        return pd.DataFrame(), set(), set(), erp_miss.copy(), ven_miss.copy()

    e = erp_miss.copy()
    v = ven_miss.copy()

    matches, used_e, used_v = [], set(), set()
    for ei, er in e.iterrows():
        if ei in used_e:
            continue
        e_inv = str(er.get("Invoice", ""))
        e_amt = round(float(er.get("Amount", 0.0)), 2)
        e_date = normalize_date(er.get("Date", "")) if "Date" in er else ""
        e_code = clean_invoice_code(e_inv)
        if not e_date:
            continue

        for vi, vr in v.iterrows():
            if vi in used_v:
                continue
            v_inv = str(vr.get("Invoice", ""))
            v_amt = round(float(vr.get("Amount", 0.0)), 2)
            v_date = normalize_date(vr.get("Date", "")) if "Date" in vr else ""
            v_code = clean_invoice_code(v_inv)
            if not v_date:
                continue

            sim = fuzzy_ratio(e_code, v_code)
            if e_date == v_date and sim >= 0.75:
                diff = abs(e_amt - v_amt)
                matches.append({
                    "ERP Invoice": e_inv,
                    "Vendor Invoice": v_inv,
                    "ERP Amount": e_amt,
                    "Vendor Amount": v_amt,
                    "Difference": round(diff, 2),
                    "Fuzzy Score": round(sim, 2),
                    "Date": e_date,
                    "Match Type": "Tier-3"
                })
                used_e.add(ei)
                used_v.add(vi)
                break

    mdf = pd.DataFrame(matches)
    rem_e = e[~e.index.isin(used_e)].copy()
    rem_v = v[~v.index.isin(used_v)].copy()
    return mdf, used_e, used_v, rem_e, rem_v

# ------- Payments detection & matching (reason + invoice text) -------
def extract_payments(erp_df, ven_df):
    pay_kw = [
        "πληρωμή", "payment", "remittance", "bank transfer",
        "transferencia", "trf", "remesa", "pago", "deposit",
        "μεταφορά", "έμβασμα", "εξόφληση", "pagado", "paid", "cobro"
    ]
    excl_kw = [
        "invoice of expenses", "expense invoice", "τιμολόγιο εξόδων",
        "διόρθωση", "correction", "reclass", "adjustment",
        "μεταφορά υπολοίπου"
    ]

    def is_payment(row, tag):
        txt = (str(row.get(f"reason_{tag}", "")) + " " + str(row.get(f"invoice_{tag}", ""))).lower()
        return any(k in txt for k in pay_kw) and not any(b in txt for b in excl_kw)

    erp_pay = erp_df[erp_df.apply(lambda r: is_payment(r, "erp"), axis=1)].copy()
    ven_pay = ven_df[ven_df.apply(lambda r: is_payment(r, "ven"), axis=1)].copy()

    # ---- helper: compute Amount column with strong fallbacks ----
    def compute_amounts(df, tag):
        if df.empty:
            return df

        # ensure numeric debit/credit columns exist
        if f"debit_{tag}" not in df.columns:
            df[f"debit_{tag}"] = 0
        if f"credit_{tag}" not in df.columns:
            df[f"credit_{tag}"] = 0

        df["Debit"]  = df[f"debit_{tag}"].apply(normalize_number)
        df["Credit"] = df[f"credit_{tag}"].apply(normalize_number)

        # Base rule: absolute difference
        base_amount = (df["Debit"] - df["Credit"]).abs().round(2)

        # If that's zero, use the larger side.
        side_max = pd.Series(
            [max(abs(d), abs(c)) for d, c in zip(df["Debit"], df["Credit"])],
            index=df.index
        ).round(2)

        # Fallback: scan any alternative amountish columns
        candidate_words = [
            "amount", "importe", "valor", "total", "document value", "net", "paid",
            "cobro", "pago", "charge", "base imponible", "importe factura", "importe neto"
        ]
        amount_like_cols = [
            c for c in df.columns
            if any(w in str(c).lower() for w in candidate_words)
            and c not in {f"debit_{tag}", f"credit_{tag}", "Debit", "Credit"}
        ]

        fallback_vals = pd.Series(0.0, index=df.index)
        for c in amount_like_cols:
            vals = df[c].apply(normalize_number).abs()
            fallback_vals = pd.concat([fallback_vals, vals], axis=1).max(axis=1)

        # Pick, in order:
        # 1) base_amount if > 0
        # 2) side_max if > 0
        # 3) fallback_vals if > 0
        df["Amount"] = base_amount
        zero_mask = df["Amount"] == 0
        df.loc[zero_mask, "Amount"] = side_max[zero_mask]
        zero_mask = df["Amount"] == 0
        df.loc[zero_mask, "Amount"] = fallback_vals[zero_mask]
        df["Amount"] = df["Amount"].abs().round(2)

        return df

    erp_pay = compute_amounts(erp_pay, "erp")
    ven_pay = compute_amounts(ven_pay, "ven")

    # ---- Match ERP ↔ Vendor payments by amount (tolerance €0.05) ----
    matched, used_v = [], set()
    for _, e in erp_pay.iterrows():
        for vi, v in ven_pay.iterrows():
            if vi in used_v:
                continue
            if abs(e["Amount"] - v["Amount"]) <= 0.05:
                matched.append({
                    "ERP Reason": e.get("reason_erp", ""),
                    "Vendor Reason": v.get("reason_ven", ""),
                    "ERP Amount": float(e["Amount"]),
                    "Vendor Amount": float(v["Amount"]),
                    "Difference": round(abs(e["Amount"] - v["Amount"]), 2)
                })
                used_v.add(vi)
                break

    pay_match = pd.DataFrame(matched)
    return erp_pay, ven_pay, pay_match

# ==================== EXCEL EXPORT =========================
def export_excel(miss_erp, miss_ven):
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Missing")

    def hdr(ws, row, color):
        for c in ws[row]:
            c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

    cur = 1
    if not miss_ven.empty:
        ws1.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=max(3, miss_ven.shape[1]))
        ws1.cell(cur, 1, "Missing in ERP").font = Font(bold=True, size=14)
        cur += 2
        for r in dataframe_to_rows(miss_ven, index=False, header=True):
            ws1.append(r)
        hdr(ws1, cur, "C62828")
        cur = ws1.max_row + 3
    if not miss_erp.empty:
        ws1.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=max(3, miss_erp.shape[1]))
        ws1.cell(cur, 1, "Missing in Vendor").font = Font(bold=True, size=14)
        cur += 2
        for r in dataframe_to_rows(miss_erp, index=False, header=True):
            ws1.append(r)
        hdr(ws1, cur, "AD1457")

    for col in ws1.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ==================== UI ==========================
st.markdown("### Upload Your Files")
uploaded_erp = st.file_uploader("ERP Export (Excel)", type=["xlsx"], key="erp")
uploaded_vendor = st.file_uploader("Vendor Statement (Excel)", type=["xlsx"], key="vendor")

if uploaded_erp and uploaded_vendor:
    try:
        erp_raw = pd.read_excel(uploaded_erp, dtype=str)
        ven_raw = pd.read_excel(uploaded_vendor, dtype=str)

        erp_df = normalize_columns(erp_raw, "erp")
        ven_df = normalize_columns(ven_raw, "ven")
        st.write("🧩 ERP columns detected:", list(erp_df.columns))
        st.write("🧩 Vendor columns detected:", list(ven_df.columns))

        with st.spinner("Analyzing invoices..."):
            # Tier-1
            tier1, miss_erp, miss_ven = match_invoices(erp_df, ven_df)

            # progressive de-dup after Tier-1
            used_erp_inv = set(tier1["ERP Invoice"].astype(str)) if not tier1.empty else set()
            used_ven_inv = set(tier1["Vendor Invoice"].astype(str)) if not tier1.empty else set()
            if not miss_erp.empty:
                miss_erp = miss_erp[~miss_erp["Invoice"].astype(str).isin(used_erp_inv)]
            if not miss_ven.empty:
                miss_ven = miss_ven[~miss_ven["Invoice"].astype(str).isin(used_ven_inv)]

            # Tier-2
            tier2, _, _, miss_erp2, miss_ven2 = tier2_match(miss_erp, miss_ven)
            if not tier2.empty:
                used_erp_inv |= set(tier2["ERP Invoice"].astype(str))
                used_ven_inv |= set(tier2["Vendor Invoice"].astype(str))
                if not miss_erp2.empty:
                    miss_erp2 = miss_erp2[
                        ~miss_erp2["Invoice"].apply(clean_invoice_code).isin(
                            [clean_invoice_code(x) for x in used_erp_inv]
                        )
                    ]
                if not miss_ven2.empty:
                    miss_ven2 = miss_ven2[
                        ~miss_ven2["Invoice"].apply(clean_invoice_code).isin(
                            [clean_invoice_code(x) for x in used_ven_inv]
                        )
                    ]

            # Tier-3
            tier3, _, _, final_erp_miss, final_ven_miss = tier3_match(miss_erp2, miss_ven2)

            if not final_erp_miss.empty:
                final_erp_miss = final_erp_miss[
                    ~final_erp_miss["Invoice"].apply(clean_invoice_code).isin(
                        [clean_invoice_code(x) for x in used_erp_inv]
                    )
                ]

            if not final_ven_miss.empty:
                final_ven_miss = final_ven_miss[
                    ~final_ven_miss["Invoice"].apply(clean_invoice_code).isin(
                        [clean_invoice_code(x) for x in used_ven_inv]
                    )
                ]

            # Payments
            erp_pay, ven_pay, pay_match = extract_payments(erp_df, ven_df)

        st.success("Reconciliation Complete!")

        # ---------- METRICS ----------
        st.markdown('<h2 class="section-title">Reconciliation Summary</h2>', unsafe_allow_html=True)
        c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
        perf = tier1[tier1["Status"] == "Perfect Match"] if not tier1.empty else pd.DataFrame()
        diff = tier1[tier1["Status"] == "Difference Match"] if not tier1.empty else pd.DataFrame()

        def safe_sum(df, col):
            return float(df[col].sum()) if not df.empty and col in df.columns else 0.0

        with c1:
            st.markdown('<div class="metric-container perfect-match">', unsafe_allow_html=True)
            st.metric("Perfect Matches", len(perf))
            st.markdown(
                f"**ERP:** {safe_sum(perf, 'ERP Amount'):,.2f}<br>"
                f"**Vendor:** {safe_sum(perf, 'Vendor Amount'):,.2f}<br>"
                f"**Diff:** {safe_sum(perf, 'Difference'):,.2f}",
                unsafe_allow_html=True
            )
            st.markdown('</div>', unsafe_allow_html=True)

        with c2:
            st.markdown('<div class="metric-container difference-match">', unsafe_allow_html=True)
            st.metric("Differences", len(diff))
            st.markdown(
                f"**ERP:** {safe_sum(diff, 'ERP Amount'):,.2f}<br>"
                f"**Vendor:** {safe_sum(diff, 'Vendor Amount'):,.2f}<br>"
                f"**Diff:** {safe_sum(diff, 'Difference'):,.2f}",
                unsafe_allow_html=True
            )
            st.markdown('</div>', unsafe_allow_html=True)

        with c3:
            st.markdown('<div class="metric-container tier2-match">', unsafe_allow_html=True)
            st.metric("Tier-2", len(tier2))
            st.markdown(
                f"**ERP:** {safe_sum(tier2, 'ERP Amount'):,.2f}<br>"
                f"**Vendor:** {safe_sum(tier2, 'Vendor Amount'):,.2f}<br>"
                f"**Diff:** {safe_sum(tier2, 'Difference'):,.2f}",
                unsafe_allow_html=True
            )
            st.markdown('</div>', unsafe_allow_html=True)

        with c4:
            st.markdown('<div class="metric-container tier3-match">', unsafe_allow_html=True)
            st.metric("Tier-3", len(tier3))
            st.markdown(
                f"**ERP:** {safe_sum(tier3, 'ERP Amount'):,.2f}<br>"
                f"**Vendor:** {safe_sum(tier3, 'Vendor Amount'):,.2f}<br>"
                f"**Diff:** {safe_sum(tier3, 'Difference'):,.2f}",
                unsafe_allow_html=True
            )
            st.markdown('</div>', unsafe_allow_html=True)

        with c5:
            st.markdown('<div class="metric-container missing-erp">', unsafe_allow_html=True)
            st.metric("Unmatched ERP", 0 if final_erp_miss.empty else len(final_erp_miss))
            st.markdown(
                f"**Total:** {final_erp_miss['Amount'].sum():,.2f}" if not final_erp_miss.empty and 'Amount' in final_erp_miss.columns else "**Total:** 0.00",
                unsafe_allow_html=True
            )
            st.markdown('</div>', unsafe_allow_html=True)

        with c6:
            st.markdown('<div class="metric-container missing-vendor">', unsafe_allow_html=True)
            st.metric("Unmatched Vendor", 0 if final_ven_miss.empty else len(final_ven_miss))
            st.markdown(
                f"**Total:** {final_ven_miss['Amount'].sum():,.2f}" if not final_ven_miss.empty and 'Amount' in final_ven_miss.columns else "**Total:** 0.00",
                unsafe_allow_html=True
            )
            st.markdown('</div>', unsafe_allow_html=True)

        with c7:
            st.markdown('<div class="metric-container payment-match">', unsafe_allow_html=True)
            if not pay_match.empty:
                total_val = pay_match[["ERP Amount", "Vendor Amount"]].apply(pd.to_numeric, errors="coerce").mean(axis=1).sum()
                st.metric("New Payment Matches (€)", f"{total_val:,.2f}")
            else:
                st.metric("New Payment Matches (€)", "0.00")
            st.markdown('</div>', unsafe_allow_html=True)

        # ---- Balance Summary Metric (C8 • Yellow • auto-detect & tolerant) ----
        with c8:
            def parse_amt(v):
                if v is None or str(v).strip() == "":
                    return np.nan
                s = str(v).replace("€", "").replace(",", ".").replace(" ", "").strip()
                s = re.sub(r"[^\d.\-]", "", s)
                try:
                    return float(s)
                except:
                    return np.nan

            # detect balance columns by name (case insensitive)
            balance_col_erp = next((c for c in erp_df.columns if "balance" in c.lower() or "saldo" in c.lower()), None)
            balance_col_ven = next((c for c in ven_df.columns if "balance" in c.lower() or "saldo" in c.lower() or "υπολ" in c.lower()), None)

            if balance_col_erp and balance_col_ven:
                erp_vals = erp_df[balance_col_erp].apply(parse_amt).dropna()
                ven_vals = ven_df[balance_col_ven].apply(parse_amt).dropna()

                if not erp_vals.empty and not ven_vals.empty:
                    erp_last = erp_vals.iloc[-1]
                    ven_last = ven_vals.iloc[-1]
                    diff_val = round(erp_last - ven_last, 2)

                    st.markdown(
                        '<div class="metric-container" style="background:#FBC02D;color:#000;font-weight:bold;">',
                        unsafe_allow_html=True
                    )
                    st.metric("Balance Difference", f"{diff_val:,.2f}")
                    st.markdown(
                        f"**ERP:** {erp_last:,.2f}<br>"
                        f"**Vendor:** {ven_last:,.2f}<br>"
                        f"**Diff:** {diff_val:,.2f}",
                        unsafe_allow_html=True
                    )
                    st.markdown('</div>', unsafe_allow_html=True)
                else:
                    st.warning("⚠️ No numeric values found in one of the Balance columns.")
            else:
                st.info("ℹ️ Could not detect Balance columns in both files.")

        st.markdown("---")

        # ---------- DISPLAY ----------
        st.markdown('<h2 class="section-title">Tier-1: Exact Matches</h2>', unsafe_allow_html=True)
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**Perfect Matches**")
            if not perf.empty:
                st.dataframe(
                    style(
                        perf[['ERP Invoice', 'Vendor Invoice', 'ERP Amount', 'Vendor Amount', 'Difference']],
                        "background:#2E7D32;color:#fff;font-weight:bold;"
                    ),
                    width="stretch"
                )
            else:
                st.info("No perfect matches.")
        with col_b:
            st.markdown("**Amount Differences**")
            if not diff.empty:
                st.dataframe(
                    style(
                        diff[['ERP Invoice', 'Vendor Invoice', 'ERP Amount', 'Vendor Amount', 'Difference']],
                        "background:#FF8F00;color:#fff;font-weight:bold;"
                    ),
                    width="stretch"
                )
            else:
                st.success("No differences.")

        st.markdown('<h2 class="section-title">Tier-2: Fuzzy + Small Amount</h2>', unsafe_allow_html=True)
        if not tier2.empty:
            st.dataframe(style(tier2, "background:#26A69A;color:#fff;font-weight:bold;"), width="stretch")
        else:
            st.info("No Tier-2 matches.")

        st.markdown('<h2 class="section-title">Tier-3: Date + Strict Fuzzy</h2>', unsafe_allow_html=True)
        if not tier3.empty:
            st.dataframe(style(tier3, "background:#7E57C2;color:#fff;font-weight:bold;"), width="stretch")
        else:
            st.info("No Tier-3 matches.")

        col_m1, col_m2 = st.columns(2)
        with col_m1:
            st.markdown('<h2 class="section-title">Missing in ERP</h2>', unsafe_allow_html=True)
            if not final_ven_miss.empty:
                st.dataframe(style(final_ven_miss, "background:#AD1457;color:#fff;font-weight:bold;"), width="stretch")
                st.error(f"{len(final_ven_miss)} vendor invoices missing – {final_ven_miss['Amount'].sum():,.2f}")
            else:
                st.success("All vendor invoices found in ERP.")
        with col_m2:
            st.markdown('<h2 class="section-title">Missing in Vendor</h2>', unsafe_allow_html=True)
            if not final_erp_miss.empty:
                st.dataframe(style(final_erp_miss, "background:#C62828;color:#fff;font-weight:bold;"), width="stretch")
                st.error(f"{len(final_erp_miss)} ERP invoices missing – {final_erp_miss['Amount'].sum():,.2f}")
            else:
                st.success("All ERP invoices found in vendor.")

        st.markdown('<h2 class="section-title">Payment Transactions</h2>', unsafe_allow_html=True)
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            st.markdown("**ERP Payments**")
            if not erp_pay.empty:
                disp = erp_pay[['reason_erp', 'Amount', 'credit_erp']].copy()
                disp.columns = ['Reason', 'Debit', 'Credit']
                st.dataframe(
                    disp.style.apply(lambda _: ['background:#4CAF50;color:#fff'] * len(_), axis=1),
                    width="stretch"
                )
                st.markdown(f"**Total:** {erp_pay['Amount'].sum():,.2f}")
            else:
                st.info("No ERP payments.")
        with col_p2:
            st.markdown("**Vendor Payments**")
            if not ven_pay.empty:
                disp = ven_pay[['reason_ven', 'debit_ven', 'credit_ven', 'Amount']].copy()
                disp.columns = ['Reason', 'Debit', 'Credit', 'Net']
                st.dataframe(
                    disp.style.apply(lambda _: ['background:#2196F3;color:#fff'] * len(_), axis=1),
                    width="stretch"
                )
                st.markdown(f"**Total:** {ven_pay['Amount'].sum():,.2f}")
            else:
                st.info("No vendor payments.")

        if not pay_match.empty:
            st.markdown("**Matched Payments**")
            st.dataframe(
                pay_match.style.apply(
                    lambda _: ['background:#004D40;color:#fff;font-weight:bold'] * len(_),
                    axis=1
                ),
                width="stretch"
            )

        # ---------- EXPORT ----------
        st.markdown('<h2 class="section-title">Download Report</h2>', unsafe_allow_html=True)
        excel_buf = export_excel(final_erp_miss, final_ven_miss)
        st.download_button(
            label="Download Full Excel Report",
            data=excel_buf,
            file_name="ReconRaptor_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Error: {e}")
        st.info("Check that your files contain columns like: **invoice**, **debit/credit**, **date**, **reason**")
