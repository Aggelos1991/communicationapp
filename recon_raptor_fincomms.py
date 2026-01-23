# --------------------------------------------------------------
# ReconRaptor v2.0 — Vendor Reconciliation Utility Functions
# Pure Python - No Streamlit dependencies
# --------------------------------------------------------------
import pandas as pd
import re
import requests
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from difflib import SequenceMatcher
import numpy as np

# ==================== FINCOMMS CONFIG ====================
FINCOMMS_API_URL = os.environ.get("FINCOMMS_API_URL", "https://fincomms-api.46-62-134-239.sslip.io")

# ====================== HELPERS ==========================
def fuzzy_ratio(a, b):
    """Calculate similarity ratio between two strings"""
    return SequenceMatcher(None, str(a), str(b)).ratio()

def normalize_number(v):
    """Convert various number formats to float"""
    if pd.isna(v) or str(v).strip() == "":
        return 0.0
    s = re.sub(r"[^\d,.\-]", "", str(v).strip())
    # Handle European vs US number formats
    if s.count(",") == 1 and s.count(".") == 1:
        if s.find(",") > s.find("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif s.count(",") == 1:
        s = s.replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "", s.count(".") - 1)
    try:
        return float(s)
    except:
        return 0.0

def normalize_date(v):
    """Normalize various date formats to YYYY-MM-DD"""
    if pd.isna(v) or str(v).strip() == "":
        return ""
    s = str(v).strip()
    for fmt in [
        "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
        "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
        "%d/%m/%y", "%d-%m-%y", "%d.%m.%y",
    ]:
        try:
            d = pd.to_datetime(s, format=fmt, errors="coerce")
            if not pd.isna(d):
                return d.strftime("%Y-%m-%d")
        except:
            continue
    # Fallback to pandas auto-detection
    d = pd.to_datetime(s, errors="coerce", dayfirst=True)
    return d.strftime("%Y-%m-%d") if not pd.isna(d) else ""

def clean_invoice_number(v):
    """
    Normalize invoice number for matching.
    INV00/0123, inv-000123, INV 123 all become '123'
    """
    if not v or pd.isna(v):
        return ""
    s = str(v).strip().upper()

    # Remove common prefixes
    prefixes = [
        r'^(INV|INVOICE|FACT|FACTURA|TIM|CN|CREDIT|NOTE|REF|DOC|NUM|NO|NR|AR|PA|PF|AB|APO|APD|VS)[\s\-_./]*',
        r'^[A-Z]{1,3}[\s\-_./]*',  # Any 1-3 letter prefix
    ]
    for prefix in prefixes:
        s = re.sub(prefix, '', s, flags=re.IGNORECASE)

    # Remove year patterns (2023, 2024, etc.) that might be in the middle
    s = re.sub(r'20[0-9]{2}[\s\-_./]*', '', s)

    # Remove all non-alphanumeric characters
    s = re.sub(r'[^A-Z0-9]', '', s)

    # Remove leading zeros
    s = s.lstrip('0')

    # If only digits remain, that's our normalized number
    # If alphanumeric, keep as is
    return s if s else "0"

def has_valid_invoice_number(inv_str):
    """Check if a string looks like a valid invoice number (not a payment reference)"""
    if not inv_str or pd.isna(inv_str):
        return False
    s = str(inv_str).strip()
    # Empty or very short
    if len(s) < 2:
        return False
    # Contains mostly numbers after cleaning
    cleaned = clean_invoice_number(s)
    return len(cleaned) >= 1 and cleaned != "0"

def normalize_columns(df, tag):
    """
    Map common column names to standardized names.
    Returns df with columns: invoice_{tag}, debit_{tag}, credit_{tag}, date_{tag}
    """
    mapping = {
        "invoice": [
            "invoice", "invoice number", "inv no", "inv", "factura", "fact",
            "numero", "document", "doc", "ref", "reference",
            "alternative document", "alt document", "alt. document",
            "voucher", "bill", "receipt"
        ],
        "debit": ["debit", "debe", "cargo", "dr", "charge"],
        "credit": ["credit", "haber", "abono", "cr", "payment"],
        "amount": ["amount", "importe", "valor", "total", "value", "sum", "net"],
        "date": ["date", "fecha", "data", "issue date", "posting date", "doc date", "document date"],
        "entity": ["entity", "company", "business unit", "bu", "cost center", "cc", "department", "dept", "organization", "org"]
    }

    rename_map = {}
    cols_lower = {c: str(c).strip().lower() for c in df.columns}

    for key, aliases in mapping.items():
        for col, low in cols_lower.items():
            if any(a in low for a in aliases):
                if key not in [v.split('_')[0] for v in rename_map.values()]:
                    rename_map[col] = f"{key}_{tag}"
                break

    out = df.rename(columns=rename_map)

    # Ensure debit/credit columns exist
    for req in ["debit", "credit"]:
        c = f"{req}_{tag}"
        if c not in out.columns:
            out[c] = 0.0

    # If we have amount but no debit/credit, use amount as debit
    if f"amount_{tag}" in out.columns:
        if out[f"debit_{tag}"].apply(normalize_number).sum() == 0:
            out[f"debit_{tag}"] = out[f"amount_{tag}"]

    # Normalize date column
    if f"date_{tag}" in out.columns:
        out[f"date_{tag}"] = out[f"date_{tag}"].apply(normalize_date)

    return out

# ==================== FINCOMMS AUTH & API ==========================
def fincomms_login(email, password):
    """Login to FinComms and get JWT token"""
    try:
        response = requests.post(
            f"{FINCOMMS_API_URL}/api/auth/login",
            json={"email": email, "password": password},
            headers={"Content-Type": "application/json"},
            timeout=30
        )
        if response.status_code == 200:
            data = response.json()
            return {"success": True, "token": data.get("token"), "user": data.get("user")}
        else:
            return {"success": False, "error": response.json().get("error", "Login failed")}
    except Exception as e:
        return {"success": False, "error": str(e)}

def send_to_fincomms(invoices_df, token, source="Missing in ERP"):
    """Send selected invoices to FinComms RECON queue"""
    if invoices_df.empty:
        return {"success": False, "error": "No invoices to send"}

    if not token:
        return {"success": False, "error": "Not logged in to FinComms. Please login first."}

    invoices = []
    for _, row in invoices_df.iterrows():
        invoice_data = {
            "invoice_number": str(row.get("Invoice", "")),
            "vendor": str(row.get("Vendor", "Unknown Vendor")),
            "entity": str(row.get("Entity", "")),
            "amount": float(row.get("Amount", 0)),
            "currency": "EUR",
            "source": "RECON"
        }
        invoices.append(invoice_data)

    try:
        response = requests.post(
            f"{FINCOMMS_API_URL}/api/invoices/bulk-import",
            json={"invoices": invoices},
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {token}"
            },
            timeout=30
        )

        if response.status_code == 201:
            return response.json()
        elif response.status_code == 401:
            return {"success": False, "error": "Session expired. Please login again."}
        else:
            return {"success": False, "error": response.text}
    except Exception as e:
        return {"success": False, "error": str(e)}

# ==================== AMOUNT CALCULATION ==========================
def calculate_amount_vendor(row, debit_col, credit_col):
    """
    Calculate amount from VENDOR statement (AP perspective).
    Vendor Debit = you owe them (positive)
    Vendor Credit = they owe you / credit note (negative)
    Negative in Debit = credit note (negative)
    """
    debit = normalize_number(row.get(debit_col, 0))
    credit = normalize_number(row.get(credit_col, 0))

    # Vendor perspective: Debit is what you owe, Credit is credit note
    # Handle negative values in either column
    if debit != 0:
        return debit  # Positive = invoice, Negative = credit note
    elif credit != 0:
        return -abs(credit)  # Credit column always represents credit note
    return 0.0

def calculate_amount_erp(row, debit_col, credit_col):
    """
    Calculate amount from YOUR ERP (AP perspective).
    ERP Credit = you owe vendor (positive AP liability)
    ERP Debit = credit note / reduces AP (negative)
    Negative in Credit = credit note (negative)
    """
    debit = normalize_number(row.get(debit_col, 0))
    credit = normalize_number(row.get(credit_col, 0))

    # AP perspective: Credit is liability (positive), Debit reduces it (negative)
    if credit != 0:
        return credit  # Positive = invoice, Negative = credit note
    elif debit != 0:
        return -abs(debit)  # Debit column always reduces AP
    return 0.0

# ==================== CONSOLIDATION ==========================
def consolidate_invoices(df, tag, amount_func):
    """
    Consolidate multiple rows with same invoice number into single net amount.
    If net = 0, the invoice is fully cancelled and excluded.
    """
    if f"invoice_{tag}" not in df.columns:
        return pd.DataFrame(), pd.DataFrame()

    df = df.copy()

    # Calculate amount for each row
    debit_col = f"debit_{tag}"
    credit_col = f"credit_{tag}"
    df["__amount"] = df.apply(lambda r: amount_func(r, debit_col, credit_col), axis=1)

    # Clean invoice numbers
    df["__inv_clean"] = df[f"invoice_{tag}"].apply(clean_invoice_number)

    # Separate rows with valid invoice numbers from payments (no valid invoice)
    has_invoice = df["__inv_clean"].apply(lambda x: x and x != "0")
    invoice_rows = df[has_invoice].copy()
    payment_rows = df[~has_invoice].copy()

    # Consolidate invoices
    consolidated = []
    for inv_clean, group in invoice_rows.groupby("__inv_clean", dropna=False):
        if not inv_clean or inv_clean == "0":
            continue

        # Sum all amounts for this invoice
        net_amount = round(group["__amount"].sum(), 2)

        # If net is 0, invoice is cancelled - skip it
        if abs(net_amount) < 0.01:
            continue

        # Take the first row as base, update amount
        base = group.iloc[0].copy()
        base["__net_amount"] = net_amount
        base["__entry_count"] = len(group)
        consolidated.append(base)

    result = pd.DataFrame(consolidated) if consolidated else pd.DataFrame()
    return result, payment_rows

# ==================== MATCHING CORE ==========================
def match_invoices(erp_df, ven_df):
    """
    Main matching function.
    Returns: matched_df, miss_erp, miss_ven
    """
    # Consolidate invoices
    erp_consolidated, erp_payments = consolidate_invoices(erp_df, "erp", calculate_amount_erp)
    ven_consolidated, ven_payments = consolidate_invoices(ven_df, "ven", calculate_amount_vendor)

    if erp_consolidated.empty and ven_consolidated.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), erp_payments, ven_payments

    # Prepare for matching
    if not erp_consolidated.empty:
        erp_consolidated["__inv_norm"] = erp_consolidated["__inv_clean"]
        erp_consolidated["__amt"] = erp_consolidated["__net_amount"].abs()

    if not ven_consolidated.empty:
        ven_consolidated["__inv_norm"] = ven_consolidated["__inv_clean"]
        ven_consolidated["__amt"] = ven_consolidated["__net_amount"].abs()

    matched = []
    used_erp = set()
    used_ven = set()

    # TIER 1: Exact invoice number match (after normalization)
    if not erp_consolidated.empty and not ven_consolidated.empty:
        for e_idx, e_row in erp_consolidated.iterrows():
            if e_idx in used_erp:
                continue

            e_inv = e_row.get("__inv_norm", "")
            e_inv_orig = str(e_row.get("invoice_erp", ""))
            e_amt = abs(float(e_row.get("__amt", 0)))
            e_entity = str(e_row.get("entity_erp", "")) if "entity_erp" in e_row else ""

            for v_idx, v_row in ven_consolidated.iterrows():
                if v_idx in used_ven:
                    continue

                v_inv = v_row.get("__inv_norm", "")
                v_inv_orig = str(v_row.get("invoice_ven", ""))
                v_amt = abs(float(v_row.get("__amt", 0)))
                v_entity = str(v_row.get("entity_ven", "")) if "entity_ven" in v_row else ""

                # Exact match on normalized invoice number
                if e_inv == v_inv and e_inv:
                    diff = abs(e_amt - v_amt)
                    status = "Perfect Match" if diff < 0.01 else "Difference Match"

                    matched.append({
                        "ERP Invoice": e_inv_orig,
                        "Vendor Invoice": v_inv_orig,
                        "ERP Amount": round(e_amt, 2),
                        "Vendor Amount": round(v_amt, 2),
                        "Difference": round(diff, 2),
                        "ERP Entity": e_entity,
                        "Vendor Entity": v_entity,
                        "Status": status,
                        "Match Type": "Tier-1"
                    })
                    used_erp.add(e_idx)
                    used_ven.add(v_idx)
                    break

    matched_df = pd.DataFrame(matched)

    # Get unmatched
    miss_erp = erp_consolidated[~erp_consolidated.index.isin(used_erp)].copy() if not erp_consolidated.empty else pd.DataFrame()
    miss_ven = ven_consolidated[~ven_consolidated.index.isin(used_ven)].copy() if not ven_consolidated.empty else pd.DataFrame()

    # Format for output
    if not miss_erp.empty:
        miss_erp = miss_erp.rename(columns={
            "invoice_erp": "Invoice",
            "__amt": "Amount",
            "date_erp": "Date",
            "entity_erp": "Entity"
        })
        keep_cols = ["Invoice", "Amount", "Date", "Entity"]
        miss_erp = miss_erp[[c for c in keep_cols if c in miss_erp.columns]].reset_index(drop=True)

    if not miss_ven.empty:
        miss_ven = miss_ven.rename(columns={
            "invoice_ven": "Invoice",
            "__amt": "Amount",
            "date_ven": "Date",
            "entity_ven": "Entity"
        })
        keep_cols = ["Invoice", "Amount", "Date", "Entity"]
        miss_ven = miss_ven[[c for c in keep_cols if c in miss_ven.columns]].reset_index(drop=True)

    return matched_df, miss_erp, miss_ven, erp_payments, ven_payments

# ==================== TIER 2 & 3 MATCHING ==========================
def tier2_match(erp_miss, ven_miss):
    """
    Tier-2: Fuzzy invoice number match (>=90%) + amount within 1 EUR
    """
    if erp_miss.empty or ven_miss.empty:
        return pd.DataFrame(), erp_miss.copy(), ven_miss.copy()

    matches = []
    used_e, used_v = set(), set()

    for ei, er in erp_miss.iterrows():
        if ei in used_e:
            continue

        e_inv = str(er.get("Invoice", ""))
        e_amt = abs(float(er.get("Amount", 0)))
        e_code = clean_invoice_number(e_inv)
        e_entity = str(er.get("Entity", "")) if "Entity" in er else ""

        best_match = None
        best_score = 0

        for vi, vr in ven_miss.iterrows():
            if vi in used_v:
                continue

            v_inv = str(vr.get("Invoice", ""))
            v_amt = abs(float(vr.get("Amount", 0)))
            v_code = clean_invoice_number(v_inv)
            v_entity = str(vr.get("Entity", "")) if "Entity" in vr else ""

            # Amount must be within 1 EUR
            if abs(e_amt - v_amt) > 1.00:
                continue

            # Calculate similarity
            if e_code in v_code or v_code in e_code:
                sim = 1.0
            else:
                sim = fuzzy_ratio(e_code, v_code)

            if sim >= 0.90 and sim > best_score:
                best_score = sim
                best_match = (vi, vr, v_inv, v_amt, sim, v_entity)

        if best_match:
            vi, vr, v_inv, v_amt, sim, v_entity = best_match
            diff = abs(e_amt - v_amt)
            matches.append({
                "ERP Invoice": e_inv,
                "Vendor Invoice": v_inv,
                "ERP Amount": round(e_amt, 2),
                "Vendor Amount": round(v_amt, 2),
                "Difference": round(diff, 2),
                "ERP Entity": e_entity,
                "Vendor Entity": v_entity,
                "Fuzzy Score": round(sim, 2),
                "Match Type": "Tier-2"
            })
            used_e.add(ei)
            used_v.add(vi)

    mdf = pd.DataFrame(matches)
    rem_e = erp_miss[~erp_miss.index.isin(used_e)].copy()
    rem_v = ven_miss[~ven_miss.index.isin(used_v)].copy()
    return mdf, rem_e, rem_v

def tier3_match(erp_miss, ven_miss):
    """
    Tier-3: Same date + fuzzy match (>=75%)
    """
    if erp_miss.empty or ven_miss.empty:
        return pd.DataFrame(), erp_miss.copy(), ven_miss.copy()

    if "Date" not in erp_miss.columns or "Date" not in ven_miss.columns:
        return pd.DataFrame(), erp_miss.copy(), ven_miss.copy()

    matches = []
    used_e, used_v = set(), set()

    for ei, er in erp_miss.iterrows():
        if ei in used_e:
            continue

        e_inv = str(er.get("Invoice", ""))
        e_amt = abs(float(er.get("Amount", 0)))
        e_date = str(er.get("Date", ""))
        e_code = clean_invoice_number(e_inv)
        e_entity = str(er.get("Entity", "")) if "Entity" in er else ""

        if not e_date:
            continue

        for vi, vr in ven_miss.iterrows():
            if vi in used_v:
                continue

            v_inv = str(vr.get("Invoice", ""))
            v_amt = abs(float(vr.get("Amount", 0)))
            v_date = str(vr.get("Date", ""))
            v_code = clean_invoice_number(v_inv)
            v_entity = str(vr.get("Entity", "")) if "Entity" in vr else ""

            if not v_date or e_date != v_date:
                continue

            sim = fuzzy_ratio(e_code, v_code)
            if sim >= 0.75:
                diff = abs(e_amt - v_amt)
                matches.append({
                    "ERP Invoice": e_inv,
                    "Vendor Invoice": v_inv,
                    "ERP Amount": round(e_amt, 2),
                    "Vendor Amount": round(v_amt, 2),
                    "Difference": round(diff, 2),
                    "ERP Entity": e_entity,
                    "Vendor Entity": v_entity,
                    "Fuzzy Score": round(sim, 2),
                    "Date": e_date,
                    "Match Type": "Tier-3"
                })
                used_e.add(ei)
                used_v.add(vi)
                break

    mdf = pd.DataFrame(matches)
    rem_e = erp_miss[~erp_miss.index.isin(used_e)].copy()
    rem_v = ven_miss[~ven_miss.index.isin(used_v)].copy()
    return mdf, rem_e, rem_v

# ==================== PAYMENT MATCHING ==========================
def match_payments(erp_payments, ven_payments):
    """
    Match payments by amount only (no invoice number).
    Tolerance: 0.05 EUR
    """
    if erp_payments.empty or ven_payments.empty:
        return pd.DataFrame()

    # Calculate amounts
    erp_payments = erp_payments.copy()
    ven_payments = ven_payments.copy()

    erp_payments["__pay_amt"] = erp_payments["__amount"].abs()
    ven_payments["__pay_amt"] = ven_payments["__amount"].abs()

    matched = []
    used_v = set()

    for _, e in erp_payments.iterrows():
        e_amt = float(e["__pay_amt"])
        e_date = str(e.get("date_erp", ""))

        for vi, v in ven_payments.iterrows():
            if vi in used_v:
                continue

            v_amt = float(v["__pay_amt"])
            v_date = str(v.get("date_ven", ""))

            if abs(e_amt - v_amt) <= 0.05:
                matched.append({
                    "ERP Amount": round(e_amt, 2),
                    "Vendor Amount": round(v_amt, 2),
                    "ERP Date": e_date,
                    "Vendor Date": v_date,
                    "Difference": round(abs(e_amt - v_amt), 2)
                })
                used_v.add(vi)
                break

    return pd.DataFrame(matched)

# ==================== EXCEL EXPORT =========================
def export_excel(miss_erp, miss_ven):
    """Export missing invoices to Excel"""
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Missing Invoices")

    def header_style(ws, row, color):
        for c in ws[row]:
            c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

    cur = 1
    if not miss_ven.empty:
        ws1.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=max(3, miss_ven.shape[1]))
        ws1.cell(cur, 1, "Missing in ERP (Vendor has, you don't)").font = Font(bold=True, size=14)
        cur += 2
        for r in dataframe_to_rows(miss_ven, index=False, header=True):
            ws1.append(r)
        header_style(ws1, cur, "C62828")
        cur = ws1.max_row + 3

    if not miss_erp.empty:
        ws1.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=max(3, miss_erp.shape[1]))
        ws1.cell(cur, 1, "Missing in Vendor (You have, vendor doesn't)").font = Font(bold=True, size=14)
        cur += 2
        for r in dataframe_to_rows(miss_erp, index=False, header=True):
            ws1.append(r)
        header_style(ws1, cur, "AD1457")

    # Auto-width columns
    for col in ws1.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==================== CLI / MAIN ==========================
if __name__ == "__main__":
    print("ReconRaptor v2.0 - Vendor Reconciliation Utility")
    print("This module provides reconciliation functions for use by the React app.")
    print("Import this module to use the matching functions programmatically.")
